#!python3
# -*- coding: utf-8 -*-

import os
import numpy
import pandas
import re
import sys
sys.path.append( os.path.join(os.path.dirname(__file__), '../lib') )
import openpyxl # for xlsx

pkeys  = ["都道府県","市区町村"]
base_sheetname = "不動産取引"
base_key = "宅地(土地と建物)_件数"
comp_wsheet_keys = {
    "SUUMO売済":["新築戸建_数"   ],
    "SUUMO在庫":["新築戸建_数"   ],
    "人口世帯" :["人口","世帯"   ],
    "平均年収" :["給与収入","資産収入"],
    "地価"     :["住居系"],
    "住宅入手法":["buy_new","buy_used","build_new"],
    "用途地域" :["住居系"],
    "生産緑地" :["面積ha"],
    "同居形態" :["親族同居"],
    "年齢"     :["30～34",     "35～39",     "40～44",     "45～49",
                 "30～34_変動","35～39_変動","40～44_変動","45～49_変動"],
    "所有形態✕構造":["戸建","集合","持家","借家"]
}


def main():
    xlsx_path = os.path.join(os.path.dirname(__file__),
                             '../../resources/15DB_SRC.XLSX')
    
    wbook = openpyxl.load_workbook(xlsx_path,
                                   read_only=True,
                                   data_only=True)
    base_shname_key = "%s::%s" % (base_sheetname,base_key)
    base_wsheet_org,head_cols = load_wsheet(wbook,base_sheetname,pkeys)

    base_wsheet = base_wsheet_org.copy()
    
    for sheet_name,comp_keys in comp_wsheet_keys.items():
        comp_wsheet,head_cols_tmp = load_wsheet(wbook,sheet_name, pkeys)
        head_cols.extend( head_cols_tmp )
        
        base_wsheet = merge_hashes(base_wsheet,comp_wsheet)

        for comp_key in comp_keys:
            comp_shname_key = "%s::%s" % (sheet_name,comp_key)
            corr_result = calc_correlation(base_wsheet,
                                           base_shname_key,
                                           comp_shname_key)
            print("%s ✕ %s = %f" %(base_shname_key,
                                   comp_shname_key,
                                   corr_result) )

    disp_merged_big_sheet(pkeys,head_cols,base_wsheet)


def disp_merged_big_sheet(pkeys,head_cols,base_wsheet):
    disp_cols = pkeys
    disp_cols.extend(head_cols)
    print( "\t".join(disp_cols) )
    
    for pkey_str in base_wsheet:
        disp_cols = []
        for head_col in head_cols:
            if head_col in base_wsheet[pkey_str]:
                disp_cols.append( str(base_wsheet[pkey_str][head_col]) )
            else:
                disp_cols.append( "" )
                
        print("%s\t%s" % (pkey_str, "\t".join(disp_cols) ) )
    
def calc_correlation(vals_tmp,base_key,comp_key):
    list_0 = []
    list_1 = []
    for pref_city, vals in vals_tmp.items():
        if not base_key in vals or \
           not comp_key in vals:
            continue
        
        list_0.append(vals[base_key])
        list_1.append(vals[comp_key])

    series_0 = pandas.Series(list_0)
    series_1 = pandas.Series(list_1)

    return series_0.corr(series_1) # 相関係数


def merge_hashes(vals_tmp, vals_tmp_2):
    for pkey in vals_tmp:
        if pkey in vals_tmp_2:
            vals_tmp[pkey].update( vals_tmp_2[pkey] )
            
    return vals_tmp
        
def load_wsheet(wbook, shname,pkeys):
    
    wsheet = wbook.get_sheet_by_name(shname)
    # ヘッダ行の取得
    col_no = 1
    head_cols = []
    while col_no <= wsheet.max_column:
        atri_key = "%s::%s" % ( shname,
                                wsheet.cell(row=1,column=col_no).value )
        head_cols.append( atri_key )
        col_no += 1

        
    ret_vals = {}
    for row_vals in wsheet.iter_rows(min_row=2):
        col_no = 0
        row_key_vals = {}
        while col_no < len(row_vals):
            atri_key = head_cols[col_no]
            row_key_vals[atri_key] = row_vals[col_no].value
            head_cols[col_no] = atri_key
            col_no += 1

        pkey_vals = []
        for pkey in pkeys:
            pkey_tmp = "%s::%s" % (shname,pkey)
            pkey_vals.append(row_key_vals[pkey_tmp])
            del row_key_vals[pkey_tmp]
        pkey_vals_str = "\t".join(pkey_vals)
        ret_vals[pkey_vals_str] = row_key_vals

    for pkey in pkeys:
        pkey_tmp = "%s::%s" % (shname,pkey)
        i = 0
        for head_col in head_cols:
            if pkey_tmp == head_col:
                del head_cols[i]
            i += 1
        
    return ret_vals, head_cols
    
if __name__ == '__main__':
    main()
