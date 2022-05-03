#!python
# -*- coding: utf-8 -*-

# refer urls are below.
# https://www.e-stat.go.jp/stat-search/files?toukei=00200521&tstat=000001049104
# https://www.e-stat.go.jp/stat-search/file-download?statInfId=000032143614&fileKind=0
# https://qiita.com/niyalist/items/d70f471c259211aa1554
# https://www.e-stat.go.jp/gis/statmap-search?page=1&type=1&toukeiCode=00200521

from bs4 import BeautifulSoup
from psycopg2  import extras # for bulk insert

import appbase
import copy
import json
import openpyxl # for xlsx, xlrd for xls
import os
import re
import tempfile
import urllib.request
from service.city       import CityService

target_host = "http://www.e-stat.go.jp"
target_tbl_name = "".join([
    "家計を主に支える者の年齢(6区分)・従業上の地位(8区分)・",
    "世帯の年間収入階級(5区分)，",
    "現住居以外の土地の所有状況(4区分)別普通世帯数―市区町村"])
download_url = \
    "https://www.e-stat.go.jp/stat-search/file-download"+ \
    "?statInfId=000032143614&fileKind=0"
master_xlsx  = "major_results_2020.xlsx"
bulk_insert_size = 20
logger = None

class PopulationCityService(appbase.AppBase):

    def __init__(self):
        global logger
        logger = self.get_logger()

    def save_tbl_rows(self, rows):
        logger.info("start")
        
        row_groups = self.__divide_rows(rows, bulk_insert_size)

        sql = """
INSERT INTO population_city 
 (pref,city,pop,pop_2015,pop_density,avg_age,
  aget_14,aget_15_64,aget_65,setai,setai_2015)
VALUES %s
"""
        with self.db_connect() as db_conn:
            with self.db_cursor(db_conn) as db_cur:

                for row_group in row_groups:
                    try:
                        # bulk insert
                        extras.execute_values(db_cur,sql,row_group)
                    except Exception as e:
                        logger.error(e)
                        logger.error(sql)
                        logger.error(row_group)
                        return False
                    
            db_conn.commit()
        return True
    

    def __divide_rows(self, org_rows, chunk_size):
        i = 0
        chunk = []
        ret_rows = []
        for org_row in org_rows:
            chunk.append( (
                org_row['pref'],
                org_row['city'],
                org_row['pop']          or None,
                org_row['pop_2015']     or None,
                org_row['pop_density']  or None,
                org_row['avg_age']      or None,
                org_row['aget_14']      or None,
                org_row['aget_15_64']   or None,
                org_row['aget_65']      or None,
                org_row['setai']        or None,
                org_row['setai_2015']   or None) )
            
            if len(chunk) >= chunk_size:
                ret_rows.append(chunk)
                chunk = []
            i += 1

        if len(chunk) > 0:
            ret_rows.append(chunk)

        return ret_rows
    
    def download_master(self):
        logger.info("start "+download_url)
        
        ret_data = []

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_xlsx_path =os.path.join(tmp_dir, master_xlsx)
            try:
                data = urllib.request.urlopen(download_url).read()
                with open(tmp_xlsx_path, mode="wb") as fh:
                    fh.write(data)
                
                wbook = openpyxl.load_workbook(tmp_xlsx_path)
                for sheetname in wbook.sheetnames:
                    logger.info("start %s %d rows" %
                                (sheetname, wbook[sheetname].max_row) )

                    tmp_ret_data = self.load_wsheet( wbook[sheetname] )
                    ret_data.extend( tmp_ret_data )
                
            except Exception as e:
                logger.error("fail", download_url)
                logger.error(e)
                return []
            
        return ret_data
        
    def load_wsheet( self, wsheet ):
        
        ret_data = []
        row_no = 8

        while row_no < wsheet.max_row :
            key_vals = \
                {"pref":       wsheet.cell(column=1, row=row_no).value,
                 "city":       wsheet.cell(column=2, row=row_no).value,
                 
                 "pop":        wsheet.cell(column=5, row=row_no).value,
                 "pop_2015":   wsheet.cell(column=8, row=row_no).value,
                 "pop_density":wsheet.cell(column=12,row=row_no).value,
                 "avg_age":    wsheet.cell(column=13,row=row_no).value,
                 "aget_14":    wsheet.cell(column=15,row=row_no).value,
                 "aget_15_64": wsheet.cell(column=16,row=row_no).value,
                 "aget_65":    wsheet.cell(column=17,row=row_no).value,
                 "setai":      wsheet.cell(column=36,row=row_no).value,
                 "setai_2015": wsheet.cell(column=39,row=row_no).value
                 }
            row_no += 1

            re_compile = re.compile("(\d+)_(.+)")
            re_result = re_compile.search(key_vals["pref"])
            if not re_result:
                continue
            key_vals["pref"] = re_result.group(2)
            
            re_result = re_compile.search(key_vals["city"])
            if not re_result:
                continue
            key_vals["city"] = re_result.group(2)

            if key_vals["pref"] == key_vals["city"]:
                continue

            for atri_key in ["pop","pop_2015","pop_density","avg_age",
                             "aget_14","aget_15_64","aget_65","setai",
                             "setai_2015"]:
                if key_vals[atri_key] == "-":
                    key_vals[atri_key] = None

            
            ret_data.append( key_vals )
            
            if row_no % 100 == 0:
                logger.info("%d %s %s" %
                            (row_no,key_vals["pref"],key_vals["city"]))

        return ret_data
