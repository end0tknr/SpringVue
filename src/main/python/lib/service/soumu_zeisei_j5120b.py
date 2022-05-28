#!python
# -*- coding: utf-8 -*-

from service.city       import CityService
from util.db import Db

import appbase
import copy
import openpyxl # for xlsx
import os
import re
import tempfile
import urllib.request

# refer urls are below.
#   https://www.soumu.go.jp/main_sosiki/jichi_zeisei/czaisei/czaisei_seido/ichiran09_20.html

download_url = "https://www.soumu.go.jp" + \
    "/main_sosiki/jichi_zeisei/czaisei/czaisei_seido/xls/J51-20-b.xlsx"
insert_cols = ["pref","city",
               "pop","salary","separate_long_capital","separate_short_capital",
               "general_stock_capital","listed_stock_capital",
               "listed_stock_dividend","future_trading_income","taxable_income"]

logger = None

class SoumuZeiseiJ5120bService(appbase.AppBase):

    def __init__(self):
        global logger
        logger = self.get_logger()

    def get_download_url(self):
        return download_url

    def del_tbl_rows(self):
        logger.info("start")
        util_db = Db()
        util_db.del_tbl_rows("soumu_zeisei_j51_20_b")

    def save_tbl_rows(self, rows):
        logger.info("start")
        util_db = Db()
        util_db.save_tbl_rows("soumu_zeisei_j51_20_b",insert_cols,rows )
        
        
    def download_src_data(self):
        download_url = self.get_download_url()
        logger.info( download_url )
        downloaded = self.download_file( download_url )

        return self.download_src_data_xlsx(downloaded)
    
    def download_src_data_xlsx(self, downloaded):
        ret_data = []
        # xlrd と異なり、openpyxl では、一度、fileに出力した上で読み込み
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_xlsx_path =os.path.join(tmp_dir, downloaded["filename"] )

            try:
                with open(tmp_xlsx_path, mode="wb") as fh:
                    fh.write( downloaded["content"] )
                
                logger.info("loading xlsx %d kbyte" %
                            ( len(downloaded["content"])/1000) )
                
                wbook = openpyxl.load_workbook(tmp_xlsx_path,
                                               read_only=True,
                                               data_only=True)
                for sheetname in wbook.sheetnames:
                    logger.info("start %s" % (sheetname) )

                    tmp_ret_data = self.load_wsheet( wbook[sheetname] )
                    ret_data.extend( tmp_ret_data )
                
            except Exception as e:
                logger.error("fail", downloaded["filename"] )
                logger.error(e)
                return []
            
        return ret_data


    def download_file(self, download_url):
        logger.info(download_url)
        try:
            res = urllib.request.urlopen(download_url)
        except Exception as e:
            logger.error(download_url)
            logger.error(e)
            return None

        content = res.read()
        filename = self.calc_download_filename(res.getheaders(),
                                               download_url )
        return {"filename":filename, "content":content}


    def calc_download_filename(self,headers, download_url):

        re_compile_1 = "^attachment;\s+"
        filename = None
        for header in headers:
            if header[0] == "Content-Disposition":
                filename = self.calc_download_filename_sub(header[1])
                break

        if filename:
            return filename

        filename = os.path.basename(download_url)
        if filename:
            return filename
        
        return None
    
                
    def load_wsheet( self, wsheet ):
        
        city_service = CityService()
        ret_data = []
        
        for row_vals in wsheet.iter_rows(min_row=3, values_only=True):
            row_vals = list(row_vals)

            if row_vals[4] != "市町村民税":
                continue

            new_info = {
                "pref"                  : row_vals[2],
                "city"                  : row_vals[3],
                "pop"                   : row_vals[5],
                "salary"                : row_vals[6],
                "separate_long_capital" : row_vals[7],
                "separate_short_capital": row_vals[8],
                "general_stock_capital" : row_vals[9],
                "listed_stock_capital"  : row_vals[10],
                "listed_stock_dividend" : row_vals[11],
                "future_trading_income" : row_vals[12],
                "taxable_income"        : row_vals[13]
            }
            ret_data.append(new_info)

        return ret_data


    def get_vals(self):
        sql = """
select
  pref, city, pop,
  salary * 1000 / pop as salary,
 (separate_long_capital +
  separate_short_capital +
  general_stock_capital +
  listed_stock_capital +
  listed_stock_dividend +
  future_trading_income ) * 1000 / pop as capital_income
from soumu_zeisei_j51_20_b
"""
        ret_data = []
        
        with self.db_connect() as db_conn:
            with self.db_cursor(db_conn) as db_cur:
                try:
                    db_cur.execute(sql)
                    
                except Exception as e:
                    logger.error(e)
                    logger.error(sql)
                    return []

                city_service = CityService()
                for ret_row in  db_cur.fetchall():
                    ret_row = dict( ret_row )
                    if not city_service.is_seirei_city(ret_row["city"]):
                        ret_data.append(ret_row)
                        continue

                    wards = city_service.get_seirei_wards(ret_row["city"])
                    for ward in wards:
                        ret_row_cp = ret_row.copy()
                        ret_row_cp["city"] = ward["city"]
                        ret_data.append(ret_row_cp)

        return ret_data
    
